#!/usr/bin/env python3
"""
Build the normalized, typed dataset for the /discovery tool.

Source of truth (both vendored under tools/data/):
  - dpdpa_personal_data_map_v3.2_all_niches.xlsx  → items, obligations, tag weights
  - dpdpa_taxonomy_master_v3.html (embedded DB)   → categories + niches (clean ids)

Why two sources: the taxonomy HTML carries the clean category/niche ids, aliases,
roles and phases; the xlsx carries the per-item detail incl. the DPDPA Obligation
Trigger column that the prototype's dpdpa-data.json had dropped. They reconcile
exactly (23 cats / 276 niches), asserted below.

Output: lib/discovery/data.generated.ts  (DO NOT EDIT BY HAND)
Run:    python3 tools/build-discovery-data.py
"""
import json, re, os, sys
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip3 install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "data", "dpdpa_personal_data_map_v3.2_all_niches.xlsx")
HTML = os.path.join(HERE, "data", "dpdpa_taxonomy_master_v3.html")
OUT  = os.path.abspath(os.path.join(HERE, "..", "lib", "discovery", "data.generated.ts"))
GOLDEN = os.path.join(HERE, "data", "niche-items.golden.json")  # raw read, for the round-trip test

# ---- Scoring model: constants, verified against Scoring_Model sheet + prototype ----
# Faithful to engine.js machine bands (continuous `score < max`); matches the
# v3.2 Scoring_Model prose within sub-point rounding.
SCORING = OrderedDict([
    ("bucketMult", {"Core": 1.0, "Operational": 1.2, "Hidden": 1.5}),
    ("q1", {"yes": 0.9, "partially": 1.0, "notsure": 1.05, "no": 1.25}),
    ("q2", {"no": 0.95, "notsure": 1.05, "yes": 1.15}),
    ("q3", {"yes": 0.9, "partially": 1.0, "notsure": 1.05, "no": 1.2}),
    ("controlCap", 1.5),
    ("modifierCap", 1.3),
    ("totalCap", 1.6),
    ("bands", [
        {"max": 35, "label": "Low"},
        {"max": 55, "label": "Moderate"},
        {"max": 80, "label": "High"},
        {"max": 9999, "label": "Critical"},
    ]),
    ("confidence", [
        {"max": 0.1, "label": "High"},
        {"max": 0.25, "label": "Medium"},
        {"max": 1.1, "label": "Low"},
    ]),
])

OBLIGATION_ORDER = [
    "Notice", "Consent", "Security safeguards",
    "Retention & erasure", "Vendor controls", "Children (verifiable consent)",
]


def extract_db(html):
    j = html.find("const DB =") + len("const DB =")
    chunk = html[j:]
    depth = 0; instr = False; esc = False; qc = ""; end = 0
    for k, ch in enumerate(chunk):
        if instr:
            if esc: esc = False
            elif ch == "\\": esc = True
            elif ch == qc: instr = False
        else:
            if ch in "\"'": instr = True; qc = ch
            elif ch == "{": depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0: end = k + 1; break
    return json.loads(chunk[:end])


def split_multi(val):
    if val is None: return []
    return [p.strip() for p in re.split(r"[;,]", str(val)) if p.strip()]


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    DB = extract_db(open(HTML, encoding="utf-8").read())

    # ---- tag library (weights) ----
    tag_lib = OrderedDict()
    it = wb["Risk_Tag_Library"].iter_rows(values_only=True)
    next(it)
    for r in it:
        if not r[0]: continue
        tag_lib[r[0]] = {"weight": int(r[3]), "meaning": r[1], "usedFor": r[2]}

    # ---- items from Data_Items_Long, deduped into ITEM_DEFS + per-niche refs ----
    di = wb["Data_Items_Long"].iter_rows(values_only=True)
    hdr = list(next(di))
    col = {name: i for i, name in enumerate(hdr)}
    defs = OrderedDict()      # signature -> index
    item_defs = []            # list of ItemDef
    niche_items = OrderedDict()  # nicheId -> [ItemRef]
    golden = OrderedDict()    # nicheId -> [full DenormItem] read straight from xlsx
    valid_obl = set(OBLIGATION_ORDER)

    for r in di:
        nid = r[col["Niche ID"]]
        if not nid: continue
        tags = split_multi(r[col["Internal Risk Tags"]])
        obls = split_multi(r[col["DPDPA Obligation Trigger"]])
        bad = [o for o in obls if o not in valid_obl]
        if bad:
            sys.exit(f"Unknown obligation token {bad} in niche {nid}")
        bad_tags = [t for t in tags if t not in tag_lib]
        if bad_tags:
            sys.exit(f"Unknown risk tag {bad_tags} in niche {nid}")
        item = {
            "item": r[col["Personal Data Item"]],
            "examples": r[col["Examples"]],
            "tags": tags,
            "precaution": r[col["Recommended Precaution"]],
            "uiGroup": r[col["UI Group"]],
            "obligations": obls,
            # v1.1 — Personal Data Map (RoPA) fields: who / why / where
            "dataSubjects": r[col["Data Subjects"]] or "",
            "processingPurposes": r[col["Processing Purposes"]] or "",
            "sources": r[col["Typical Sources / Systems"]] or "",
        }
        sig = json.dumps(item, sort_keys=True, ensure_ascii=False)
        if sig not in defs:
            defs[sig] = len(item_defs)
            item_defs.append(item)
        seq = int(r[col["Item Sequence"]])
        bucket = r[col["Bucket"]]
        defsel = str(r[col["Default Selected"]]).strip().upper() == "TRUE"
        niche_items.setdefault(nid, []).append({"ref": defs[sig], "seq": seq, "bucket": bucket, "def": defsel})
        # golden = the shape denormNiche() must reproduce (ItemDef fields + seq/bucket/def)
        golden.setdefault(nid, []).append({**item, "seq": seq, "bucket": bucket, "def": defsel})

    # ---- reconcile xlsx <-> taxonomy DB ----
    db_niche_ids = {n["id"] for n in DB["niches"]}
    xl_niche_ids = set(niche_items)
    assert db_niche_ids == xl_niche_ids, (
        f"niche id mismatch: only-db={sorted(db_niche_ids-xl_niche_ids)[:5]} "
        f"only-xlsx={sorted(xl_niche_ids-db_niche_ids)[:5]}")
    assert len(DB["categories"]) == 23, f"expected 23 categories, got {len(DB['categories'])}"
    assert len(DB["niches"]) == 276, f"expected 276 niches, got {len(DB['niches'])}"

    # ---- emit TS ----
    def ts(x):
        return json.dumps(x, ensure_ascii=False, separators=(",", ":"))

    total_items = sum(len(v) for v in niche_items.values())
    lines = []
    lines.append("// AUTO-GENERATED by tools/build-discovery-data.py — DO NOT EDIT.")
    lines.append(f"// Source: dpdpa_personal_data_map_v3.2 (xlsx) + dpdpa_taxonomy_master_v3 (DB v{DB.get('version')}).")
    lines.append(f"// {len(DB['categories'])} categories · {len(DB['niches'])} niches · "
                 f"{total_items} item rows · {len(item_defs)} unique item defs · {len(tag_lib)} risk tags.")
    lines.append('import type { Category, Niche, ItemDef, ItemRef, TagLib, Scoring } from "./types";')
    lines.append("")
    lines.append(f"export const VERSION = {ts(str(DB.get('version')))};")
    lines.append(f"export const ROLES: Record<string,string> = {ts(DB.get('roles', {}))};")
    lines.append(f"export const PHASES: Record<string,string> = {ts(DB.get('phases', {}))};")
    lines.append(f"export const CATEGORIES: Category[] = {ts(DB['categories'])};")
    lines.append(f"export const NICHES: Niche[] = {ts(DB['niches'])};")
    lines.append(f"export const TAG_LIB: TagLib = {ts(tag_lib)};")
    lines.append(f"export const SCORING: Scoring = {ts(SCORING)};")
    lines.append(f"export const ITEM_DEFS: ItemDef[] = {ts(item_defs)};")
    lines.append(f"export const NICHE_ITEMS: Record<string, ItemRef[]> = {ts(niche_items)};")
    lines.append("")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    with open(GOLDEN, "w", encoding="utf-8") as f:
        json.dump(golden, f, ensure_ascii=False, separators=(",", ":"))

    print(f"✓ wrote {os.path.relpath(OUT)}")
    print(f"✓ wrote {os.path.relpath(GOLDEN)} (golden fixture, {len(golden)} niches)")
    print(f"  {len(DB['categories'])} categories · {len(DB['niches'])} niches · "
          f"{total_items} items → {len(item_defs)} unique defs · {len(tag_lib)} tags")


if __name__ == "__main__":
    main()
